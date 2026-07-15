#!/usr/bin/env python3
"""
Python tool for enriching vendor spreadsheets with missing product/service info
from web sources.

The scraper is intentionally conservative:
- it preserves every original row and column
- it only fills blank product cells unless --overwrite is passed
- it writes source/status columns so scraped values can be checked later
- it does not require paid APIs
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote_plus, unquote, urlparse
from urllib.request import Request, urlopen


DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
)

PROJECT_DESCRIPTION = (
    "Python tool for enriching vendor spreadsheets with missing product/service "
    "info from web sources."
)

FREE_EMAIL_DOMAINS = {
    "aol.com",
    "gmail.com",
    "hotmail.com",
    "icloud.com",
    "live.com",
    "me.com",
    "msn.com",
    "outlook.com",
    "roadrunner.com",
    "yahoo.com",
}

SKIP_COMPANY_RE = re.compile(
    r"\b(click|confirm|choose|push|press|success|select update|enter:|settings|"
    r"to reconnect|to update|if it asks|if you use|download/|http[s]?://)\b",
    re.I,
)

URL_RE = re.compile(r"https?://[^\s<>)\"']+", re.I)
EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@([A-Z0-9.-]+\.[A-Z]{2,})\b", re.I)
PHONE_RE = re.compile(r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}")

PRODUCT_HEADER_HINTS = (
    "product",
    "products",
    "item",
    "items",
    "service",
    "services",
    "offering",
    "category",
    "description",
)

COMPANY_HEADER_HINTS = (
    "company",
    "vendor",
    "supplier",
    "contact company",
    "contact",
    "business",
    "name",
)

SERVICE_KEYWORDS: Sequence[Tuple[str, str]] = (
    (r"\bpayroll\b", "payroll services"),
    (r"\bplumb(?:er|ing)?\b", "plumbing services"),
    (r"\block\s*smith\b|\blocksmith\b", "locksmith services"),
    (r"\bglass\b", "glass repair and replacement"),
    (r"\bappliance\b", "appliance repair"),
    (r"\bduct\b", "dryer duct cleaning"),
    (r"\bair conditioning\b|\bhvac\b", "air conditioning and HVAC services"),
    (r"\balarms?\b|\bsecurity\b", "security alarm services"),
    (r"\bmanaged it\b|\bit support\b|\binformation technology\b|itinc|mavenit", "IT support services"),
    (r"\bbeauty\b|\bcosmoprof\b|\bisp-beauty\b", "beauty products"),
    (r"\bleaf and flower\b", "hair care products"),
    (r"\bhsa\b", "HSA signup/support services"),
    (r"\bsignup\b|\bregistration\b", "registration/signup services"),
    (r"\bad america\b|\badvertising\b|\bprint(?:ing)?\b", "advertising and print services"),
)

PAGE_TEXT_KEYWORDS = (
    "products",
    "product",
    "services",
    "service",
    "solutions",
    "offer",
    "specialize",
    "repair",
    "installation",
    "maintenance",
    "support",
)


@dataclass
class ScrapeResult:
    products: str = ""
    source_url: str = ""
    confidence: str = "none"
    status: str = "not_found"
    notes: str = ""


class PageTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.title_parts: List[str] = []
        self.meta_descriptions: List[str] = []
        self.headings: List[str] = []
        self.links: List[Tuple[str, str]] = []
        self.text_parts: List[str] = []
        self._tag_stack: List[str] = []
        self._current_link: Optional[str] = None
        self._current_link_text: List[str] = []

    def handle_starttag(self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]) -> None:
        attr = {k.lower(): v or "" for k, v in attrs}
        tag = tag.lower()
        self._tag_stack.append(tag)

        if tag == "meta":
            name = (attr.get("name") or attr.get("property") or "").lower()
            if name in {"description", "og:description", "twitter:description"}:
                content = clean_text(attr.get("content", ""))
                if content:
                    self.meta_descriptions.append(content)

        if tag == "a":
            href = attr.get("href", "")
            if href:
                self._current_link = href
                self._current_link_text = []

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "a" and self._current_link:
            label = clean_text(" ".join(self._current_link_text))
            self.links.append((self._current_link, label))
            self._current_link = None
            self._current_link_text = []

        for i in range(len(self._tag_stack) - 1, -1, -1):
            if self._tag_stack[i] == tag:
                del self._tag_stack[i:]
                break

    def handle_data(self, data: str) -> None:
        text = clean_text(data)
        if not text:
            return

        current = self._tag_stack[-1] if self._tag_stack else ""
        if current == "title":
            self.title_parts.append(text)
        elif current in {"h1", "h2", "h3"}:
            self.headings.append(text)

        if current not in {"script", "style", "noscript"}:
            self.text_parts.append(text)

        if self._current_link is not None:
            self._current_link_text.append(text)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = html.unescape(str(value)).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def unique_headers(headers: Sequence[str], width: int) -> List[str]:
    cleaned: List[str] = []
    seen: Dict[str, int] = {}
    for idx in range(width):
        raw = clean_text(headers[idx] if idx < len(headers) else "")
        name = raw or f"Extra {idx + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} {seen[name]}"
        else:
            seen[name] = 1
        cleaned.append(name)
    return cleaned


def trim_empty_tail(values: Sequence[object]) -> List[str]:
    row = [clean_text(v) for v in values]
    while row and not row[-1]:
        row.pop()
    return row


def rows_from_pdf(path: Path) -> List[Dict[str, str]]:
    try:
        import pdfplumber  # type: ignore
    except ImportError as exc:
        raise RuntimeError("PDF input requires pdfplumber. Run: pip install -r requirements.txt") from exc

    output: List[Dict[str, str]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            for table_number, table in enumerate(tables, start=1):
                if not table:
                    continue
                width = max(len(r) for r in table)
                headers = unique_headers(table[0], width)
                for source_row_number, raw_row in enumerate(table[1:], start=2):
                    row_values = list(raw_row) + [""] * (width - len(raw_row))
                    row = {headers[i]: clean_text(row_values[i]) for i in range(width)}
                    if not any(row.values()):
                        continue
                    row["_source_file"] = str(path)
                    row["_source_sheet"] = f"page {page_number}, table {table_number}"
                    row["_source_row"] = str(source_row_number)
                    output.append(row)
    return output


def rows_from_xlsx(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Excel input requires openpyxl. Run: pip install -r requirements.txt") from exc

    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    header_index = find_header_row(raw_rows)
    if header_index is None:
        return []

    width = max(len(r) for r in raw_rows[header_index:])
    headers = unique_headers(raw_rows[header_index], width)
    output: List[Dict[str, str]] = []
    for excel_row_number, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        row_values = list(raw_row) + [""] * (width - len(raw_row))
        row = {headers[i]: clean_text(row_values[i]) for i in range(width)}
        if not any(row.values()):
            continue
        row["_source_file"] = str(path)
        row["_source_sheet"] = sheet.title
        row["_source_row"] = str(excel_row_number)
        output.append(row)
    return output


def rows_from_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        output: List[Dict[str, str]] = []
        for row_number, row in enumerate(reader, start=2):
            cleaned = {clean_text(k): clean_text(v) for k, v in row.items() if k is not None}
            if not any(cleaned.values()):
                continue
            cleaned["_source_file"] = str(path)
            cleaned["_source_sheet"] = ""
            cleaned["_source_row"] = str(row_number)
            output.append(cleaned)
    return output


def find_header_row(raw_rows: Sequence[Sequence[object]]) -> Optional[int]:
    best_idx = None
    best_score = 0
    for idx, row in enumerate(raw_rows[:25]):
        cells = [clean_text(v).lower() for v in row]
        if not any(cells):
            continue
        score = 0
        for cell in cells:
            if any(hint in cell for hint in COMPANY_HEADER_HINTS + PRODUCT_HEADER_HINTS):
                score += 2
            elif cell:
                score += 1
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def read_rows(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return rows_from_pdf(path)
    if suffix in {".xlsx", ".xlsm"}:
        return rows_from_xlsx(path, sheet_name)
    if suffix == ".csv":
        return rows_from_csv(path)
    raise ValueError(f"Unsupported input type: {path.suffix}")


def detect_column(headers: Iterable[str], hints: Sequence[str], blocked: Sequence[str] = ()) -> Optional[str]:
    candidates = list(headers)
    normalized = [(h, clean_text(h).lower()) for h in candidates]
    for header, lower in normalized:
        if any(block in lower for block in blocked):
            continue
        if any(hint == lower for hint in hints):
            return header
    for header, lower in normalized:
        if any(block in lower for block in blocked):
            continue
        if any(hint in lower for hint in hints):
            return header
    return None


def should_skip_company(company: str) -> bool:
    company = clean_text(company)
    if not company:
        return True
    if len(company) < 2:
        return True
    if SKIP_COMPANY_RE.search(company):
        return True
    if URL_RE.search(company):
        return True
    return False


def row_blob(row: Dict[str, str]) -> str:
    return " ".join(clean_text(v) for k, v in row.items() if not k.startswith("_"))


def extract_urls(text: str) -> List[str]:
    urls = []
    for match in URL_RE.findall(text):
        url = match.rstrip(".,;]")
        urls.append(url)
    return dedupe(urls)


def domains_from_email(text: str) -> List[str]:
    domains = []
    for domain in EMAIL_RE.findall(text):
        domain = domain.lower().strip(".")
        if domain and domain not in FREE_EMAIL_DOMAINS:
            domains.append(domain)
    return dedupe(domains)


def dedupe(values: Iterable[str]) -> List[str]:
    seen = set()
    output = []
    for value in values:
        cleaned = clean_text(value)
        key = cleaned.lower()
        if cleaned and key not in seen:
            seen.add(key)
            output.append(cleaned)
    return output


def is_bad_url(url: str) -> bool:
    parsed = urlparse(url)
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()
    if not parsed.scheme.startswith("http"):
        return True
    if any(site in host for site in ("facebook.com", "instagram.com", "linkedin.com", "yelp.com")):
        return True
    if any(part in host for part in ("google.", "bing.", "duckduckgo.com")):
        return True
    if any(word in path for word in ("/login", "/signin", "/account", "/cart")):
        return True
    return False


def fetch_url(url: str, timeout: float = 12.0) -> Tuple[str, str]:
    request = Request(url, headers={"User-Agent": DEFAULT_USER_AGENT})
    with urlopen(request, timeout=timeout) as response:
        content_type = response.headers.get("content-type", "")
        charset = response.headers.get_content_charset() or "utf-8"
        body = response.read(700_000)
    if "text/html" not in content_type and "application/xhtml" not in content_type and content_type:
        return "", content_type
    return body.decode(charset, errors="replace"), content_type


def parse_page(html_text: str) -> PageTextParser:
    parser = PageTextParser()
    parser.feed(html_text)
    return parser


def duckduckgo_search(query: str, max_results: int = 5) -> List[str]:
    search_url = "https://duckduckgo.com/html/?q=" + quote_plus(query)
    try:
        html_text, _ = fetch_url(search_url)
    except Exception:
        return []

    parser = parse_page(html_text)
    urls: List[str] = []
    for href, _label in parser.links:
        href = html.unescape(href)
        if href.startswith("//duckduckgo.com/l/?"):
            href = "https:" + href
        if "duckduckgo.com/l/?" in href:
            parsed = urlparse(href)
            qs = parse_qs(parsed.query)
            if qs.get("uddg"):
                href = unquote(qs["uddg"][0])
        if href.startswith("http") and not is_bad_url(href):
            urls.append(href)
    return dedupe(urls)[:max_results]


def infer_from_text(*parts: str) -> Optional[str]:
    lower = " ".join(parts).lower().replace("&", " and ")
    for pattern, product in SERVICE_KEYWORDS:
        if re.search(pattern, lower):
            return product
    return None


def is_generic_product_guess(products: str) -> bool:
    lower = products.lower()
    words = re.findall(r"[a-z0-9]+", lower)
    if len(words) <= 4 and not any(keyword in lower for keyword in PAGE_TEXT_KEYWORDS):
        return True
    generic_phrases = {
        "home",
        "about",
        "contact",
        "metro detroit businesses",
        "learn more",
    }
    return lower.strip(" .,:;") in generic_phrases


def score_sentence(sentence: str) -> int:
    lower = sentence.lower()
    score = 0
    for keyword in PAGE_TEXT_KEYWORDS:
        if keyword in lower:
            score += 2
    if 30 <= len(sentence) <= 220:
        score += 1
    if PHONE_RE.search(sentence) or EMAIL_RE.search(sentence):
        score -= 2
    if len(sentence) > 280:
        score -= 2
    return score


def split_sentences(text: str) -> List[str]:
    text = re.sub(r"\s+", " ", text)
    rough = re.split(r"(?<=[.!?])\s+|\s+[|]\s+|\s+[-]\s+", text)
    return [clean_text(part) for part in rough if clean_text(part)]


def products_from_page(company: str, parser: PageTextParser) -> str:
    pieces: List[str] = []
    pieces.extend(parser.meta_descriptions[:2])
    pieces.extend(parser.headings[:8])

    page_text = " ".join(parser.text_parts[:600])
    sentences = split_sentences(page_text)
    scored = sorted(
        ((score_sentence(sentence), sentence) for sentence in sentences),
        reverse=True,
        key=lambda item: (item[0], len(item[1])),
    )
    for score, sentence in scored:
        if score <= 0:
            continue
        if company.lower() in sentence.lower() and len(sentence) < 35:
            continue
        pieces.append(sentence)
        if len(pieces) >= 8:
            break

    return summarize_products(company, pieces)


def summarize_products(company: str, snippets: Sequence[str]) -> str:
    candidates: List[str] = []
    company_words = set(re.findall(r"[a-z0-9]+", company.lower()))

    for snippet in snippets:
        snippet = clean_text(snippet)
        if not snippet:
            continue

        lowered = snippet.lower()
        if len(snippet) > 240:
            snippet = snippet[:237].rsplit(" ", 1)[0] + "..."
        if PHONE_RE.search(snippet) or "privacy policy" in lowered or "terms of" in lowered:
            continue

        phrases = []
        for pattern in (
            r"(?:products?|services?|solutions?)\s+(?:include|including|are|for)\s+([^.;:]{8,120})",
            r"(?:we offer|offering|specializ(?:e|ing) in|provid(?:e|ing))\s+([^.;:]{8,120})",
            r"([^.;:]{8,120}\s+(?:services?|products?|solutions?|repair|installation|maintenance))",
        ):
            for match in re.findall(pattern, snippet, flags=re.I):
                phrase = clean_text(match)
                if phrase:
                    phrases.append(phrase)

        if phrases:
            candidates.extend(phrases)
        elif any(keyword in lowered for keyword in PAGE_TEXT_KEYWORDS):
            candidates.append(snippet)

    cleaned_candidates: List[str] = []
    for candidate in candidates:
        words = re.findall(r"[a-z0-9]+", candidate.lower())
        if len(set(words) - company_words) < 2:
            continue
        candidate = re.sub(r"\b(home|about us|contact us|learn more)\b", "", candidate, flags=re.I)
        candidate = clean_text(candidate.strip(" -:;,"))
        if candidate:
            cleaned_candidates.append(candidate)

    unique = dedupe(cleaned_candidates)
    if not unique:
        return ""

    first = unique[0]
    first = re.sub(r"^(our|we|and)\s+", "", first, flags=re.I)
    return first[:180].rstrip(" ,;:-")


def candidate_urls_for_row(company: str, row: Dict[str, str], search_results: int) -> List[str]:
    blob = row_blob(row)
    urls = [u for u in extract_urls(blob) if not is_bad_url(u)]

    for domain in domains_from_email(blob):
        urls.append("https://" + domain)
        urls.append("https://www." + domain)

    if search_results:
        query = f"{company} products services official website"
        urls.extend(duckduckgo_search(query, max_results=search_results))

    return dedupe(urls)


def scrape_products_for_row(
    company: str,
    row: Dict[str, str],
    search_results: int,
    sleep_seconds: float,
    cache: Dict[str, ScrapeResult],
    no_web: bool,
) -> ScrapeResult:
    heuristic = infer_from_text(company, row_blob(row))

    if no_web:
        if heuristic:
            return ScrapeResult(
                products=heuristic,
                confidence="low",
                status="filled_from_company_name",
                notes="Web scraping disabled; inferred from company/contact name.",
            )
        return ScrapeResult(status="not_found", notes="Web scraping disabled.")

    cache_key = company.lower() + "|" + row_blob(row).lower()
    if cache_key in cache:
        return cache[cache_key]

    errors: List[str] = []
    for url in candidate_urls_for_row(company, row, search_results):
        if is_bad_url(url):
            continue
        try:
            if sleep_seconds:
                time.sleep(sleep_seconds)
            html_text, _content_type = fetch_url(url)
            if not html_text:
                continue
            parser = parse_page(html_text)
            products = products_from_page(company, parser)
            if products:
                if heuristic and is_generic_product_guess(products):
                    products = heuristic
                result = ScrapeResult(
                    products=products,
                    source_url=url,
                    confidence="medium",
                    status="filled_from_web",
                )
                cache[cache_key] = result
                return result
        except (HTTPError, URLError, TimeoutError, OSError) as exc:
            errors.append(f"{url}: {exc.__class__.__name__}")
            continue

    if heuristic:
        result = ScrapeResult(
            products=heuristic,
            confidence="low",
            status="filled_from_company_name",
            notes="No useful web result; inferred from company/contact name.",
        )
        cache[cache_key] = result
        return result

    result = ScrapeResult(
        status="not_found",
        notes="; ".join(errors[:3]) if errors else "No usable source found.",
    )
    cache[cache_key] = result
    return result


def ordered_headers(rows: Sequence[Dict[str, str]], preferred: Sequence[str]) -> List[str]:
    headers: List[str] = []
    for header in preferred:
        if header not in headers:
            headers.append(header)
    for row in rows:
        for header in row:
            if header not in headers:
                headers.append(header)
    return headers


def write_csv(rows: Sequence[Dict[str, str]], headers: Sequence[str], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(rows: Sequence[Dict[str, str]], headers: Sequence[str], output_path: Path) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Enriched Products"
    sheet.append(list(headers))
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for cell in sheet.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            max_len = max(max_len, len(clean_text(cell[0])))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 55)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def load_cache(path: Optional[Path]) -> Dict[str, ScrapeResult]:
    if not path or not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        raw = json.load(handle)
    return {k: ScrapeResult(**v) for k, v in raw.items()}


def save_cache(path: Optional[Path], cache: Dict[str, ScrapeResult]) -> None:
    if not path:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump({k: vars(v) for k, v in cache.items()}, handle, indent=2, sort_keys=True)


def enrich_rows(args: argparse.Namespace) -> Tuple[List[Dict[str, str]], List[str]]:
    input_path = Path(args.input).expanduser().resolve()
    rows = read_rows(input_path, sheet_name=args.sheet)
    if not rows:
        raise RuntimeError(f"No rows found in {input_path}")

    all_headers = ordered_headers(rows, [])
    company_col = args.company_column or detect_column(all_headers, COMPANY_HEADER_HINTS, blocked=("phone", "email"))
    if not company_col:
        raise RuntimeError("Could not detect a company/vendor column. Pass --company-column.")

    product_col = args.product_column or detect_column(all_headers, PRODUCT_HEADER_HINTS)
    if not product_col:
        product_col = "Products"

    source_col = args.source_column
    confidence_col = args.confidence_column
    status_col = args.status_column
    notes_col = args.notes_column

    cache = load_cache(Path(args.cache).expanduser().resolve() if args.cache else None)

    enriched: List[Dict[str, str]] = []
    processed = 0
    filled = 0
    for row in rows:
        row = dict(row)
        row.setdefault(product_col, "")
        row.setdefault(source_col, "")
        row.setdefault(confidence_col, "")
        row.setdefault(status_col, "")
        row.setdefault(notes_col, "")

        company = clean_text(row.get(company_col, ""))
        has_product = bool(clean_text(row.get(product_col, "")))
        should_fill = args.overwrite or not has_product

        if not should_fill:
            row[status_col] = row[status_col] or "already_had_product"
            enriched.append(row)
            continue

        if should_skip_company(company):
            row[status_col] = row[status_col] or "skipped"
            row[notes_col] = row[notes_col] or "No vendor/company name to scrape."
            enriched.append(row)
            continue

        if args.limit and processed >= args.limit:
            row[status_col] = row[status_col] or "not_processed_limit"
            enriched.append(row)
            continue

        processed += 1
        result = scrape_products_for_row(
            company=company,
            row=row,
            search_results=args.search_results,
            sleep_seconds=args.sleep,
            cache=cache,
            no_web=args.no_web,
        )
        row[status_col] = result.status
        row[confidence_col] = result.confidence
        row[source_col] = result.source_url
        row[notes_col] = result.notes
        if result.products:
            row[product_col] = result.products
            filled += 1
        enriched.append(row)

    save_cache(Path(args.cache).expanduser().resolve() if args.cache else None, cache)

    preferred_headers = [
        company_col,
        product_col,
        source_col,
        confidence_col,
        status_col,
        notes_col,
    ]
    headers = ordered_headers(enriched, preferred_headers)
    print(f"Read {len(rows)} rows from {input_path}")
    print(f"Processed {processed} rows; filled {filled} product cells")
    return enriched, headers


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=PROJECT_DESCRIPTION)
    parser.add_argument("input", help="Input PDF, XLSX, or CSV file.")
    parser.add_argument("-o", "--output", default="output/enriched_products.csv", help="Output CSV path.")
    parser.add_argument("--xlsx-output", default="", help="Optional Excel output path.")
    parser.add_argument("--sheet", default=None, help="Worksheet name for XLSX input. Defaults to active sheet.")
    parser.add_argument("--company-column", default="", help="Column containing vendor/company names.")
    parser.add_argument("--product-column", default="", help="Column to fill. Created if omitted and not detected.")
    parser.add_argument("--source-column", default="Product Source URL")
    parser.add_argument("--confidence-column", default="Product Confidence")
    parser.add_argument("--status-column", default="Scrape Status")
    parser.add_argument("--notes-column", default="Scrape Notes")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing product values.")
    parser.add_argument("--limit", type=int, default=0, help="Only process this many missing rows; 0 means all.")
    parser.add_argument("--search-results", type=int, default=4, help="Search result URLs to try per company.")
    parser.add_argument("--sleep", type=float, default=0.5, help="Delay between HTTP requests.")
    parser.add_argument("--cache", default=".scrape_cache.json", help="JSON cache path. Empty string disables cache.")
    parser.add_argument("--no-web", action="store_true", help="Do not make web requests; use company-name inference only.")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        rows, headers = enrich_rows(args)
        output_path = Path(args.output).expanduser().resolve()
        write_csv(rows, headers, output_path)
        print(f"Wrote CSV: {output_path}")

        if args.xlsx_output:
            xlsx_path = Path(args.xlsx_output).expanduser().resolve()
            write_xlsx(rows, headers, xlsx_path)
            print(f"Wrote Excel workbook: {xlsx_path}")
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
